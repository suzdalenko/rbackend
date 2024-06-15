import os
import threading
import time
import openpyxl
from django.db import connection
from pathlib import Path
from openpyxl.styles import Alignment, PatternFill, Font, Color
from myapp.models import CollectionLines
from myapp.myresponse import DiscoJsonResponse, UserLoginCorrectry, get_current_file_directory


class Excel:

    def basic_report(request):
        # if(UserLoginCorrectry(request)):
            userId       = request.GET.get("user_id")    
            collectionId = request.GET.get("collection_id")
            disctinct_truck = "SELECT DISTINCT truck FROM colectionlines WHERE colection_id = "+str(collectionId)+" AND truck > 0 ORDER BY truck"
            cursor = connection.cursor()
            cursor.execute(disctinct_truck)
            disctinct_truck = cursor.fetchall()
            cursor.close()

            wb = openpyxl.Workbook()
            firstTemlate = 0
            ws = wb.active

            for trA in disctinct_truck:
                try:
                    lines = CollectionLines.objects.filter(colection_id=collectionId, truck=trA[0]).order_by("by_order")   
                    truckName = lines[0].truck_name
                    if truckName == None:
                        truckName = 'Camion'          
                except:
                    lines = []
                    truckName = 'Camion'
                titleSheetA = str(trA[0])+'. '+truckName

                if firstTemlate == 0:
                   ws.title = titleSheetA
                else:
                    ws = wb.create_sheet(titleSheetA)

                # ws.merged_cells("A1:E1")
                ws['A1'] = titleSheetA
                ws.append(["", "", "", "", ""])
                ws.append(["Id Pedido", "Cliente", "Cuidad", "Paletas", "Kilos", "Fecha Entrega"])

                paletA = 0
                kilosA = 0
                for l in lines:
                    ws.append([l.order_id, l.client_name, l.city, l.palets, l.kilos, l.delivery_date])
                    paletA += l.palets
                    kilosA += l.kilos
                    print(l)
                ws.append(["", "", "", paletA, kilosA, ""])

                firstTemlate += 1


                dims = {}
                letter = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
                for row in ws.rows:
                    for cell in row:
                        if cell.value:
                          dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))    
                for col, value in dims.items():
                    ws.column_dimensions[letter[col]].width = value + 11     
                   
            urlDirection = get_current_file_directory(request)
            urlFile = urlDirection+'excel_'+str(collectionId)+'.xlsx'
            wb.save(urlFile)
            urlExcel = 'static/'+'excel_'+str(collectionId)+'.xlsx'

            tA = threading.Thread(target=Excel.doCrawl, args=[urlFile])
            tA.setDaemon(True)
            tA.start()

            return DiscoJsonResponse({"res":urlExcel})
    

    def doCrawl(urlFile):
        time.sleep(111)
        try:
            pass
            # os.remove(urlFile)
        except:
            pass
        